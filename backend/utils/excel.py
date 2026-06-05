# backend/utils/excel.py
from __future__ import annotations

import io
import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# === Rutas base ===
ROOT_BACKEND = Path(__file__).resolve().parents[1]
TEMPLATES_DIR = ROOT_BACKEND / "data" / "templates"
DEFAULT_TEMPLATE_PATH = TEMPLATES_DIR / "casos_template.xlsx"

# === Helpers ENV seguros ===
def _env_bool(name: str, default: bool = False) -> bool:
    val = os.getenv(name, str(default)).strip().lower()
    return val in {"1", "true", "t", "yes", "y", "on"}

def _env_int(name: str, default: int = 0) -> int:
    try:
        return int(os.getenv(name, str(default)).strip() or default)
    except Exception:
        return default

# === Config por .env (todas opcionales) ===
LEAVE_VALIDACION_EMPTY   = _env_bool("EXCEL_LEAVE_VALIDACION_EMPTY", True)
ID_STYLE                 = os.getenv("EXCEL_ID_STYLE", "number").strip().lower()  # "number" | "cp"
FILL_FROM_TEMPLATE_BOXES = _env_bool("EXCEL_FILL_FROM_TEMPLATE_BOXES", True)
OBS_CLASSIFY             = _env_bool("EXCEL_OBS_CLASSIFY", True)

# Funcionalidad: inferencia / override y fuerza de escritura
EXCEL_FUNC_OVERRIDE = os.getenv("EXCEL_FUNC_OVERRIDE", "").strip()
EXCEL_FUNC_DEFAULT  = os.getenv("EXCEL_FUNC_DEFAULT",  "").strip()
EXCEL_FUNC_FORCE    = _env_bool("EXCEL_FUNC_FORCE", False)   # por defecto NO fuerza
EXCEL_FUNC_COL      = _env_int("EXCEL_FUNC_COL", 4)          # fallback col index (D=4) si no encuentra header

# === Columnas fuente (input del generador) ===
SOURCE_HEADERS = [
    "Caso de Prueba",
    "Funcionalidad",
    "Objetivo",
    "Paso a Paso",
    "Resultado Esperado",
    "Precondiciones",
    "Prioridad",
    "Datos de Prueba",
    "Observaciones",
]

# === Encabezados del TEMPLATE ===
TEMPLATE_HEADERS = [
    "Id. Caso de Prueba",
    "Objetivo de la prueba",
    "Funcionalidad",
    "Resultado Esperado",
    "Validación del resultado",
    "Observaciones",
]

# ---------------- Helpers de normalización ----------------
def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", " ", s).strip().lower()
    return s

def _sanitize(text: str) -> str:
    if text is None:
        return ""
    t = str(text)
    t = re.sub(r"<br\s*/?>", "\n", t, flags=re.IGNORECASE)
    t = re.sub(r"<[^>]+>", "", t)
    t = re.sub(r"[ \t]+\n", "\n", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t

def _title_case_es(s: str) -> str:
    s = s.strip()
    return s[:1].upper() + s[1:] if s else s

# ---------- Resolver ruta del template (FALTABA ESTA FUNCIÓN) ----------
def resolve_template_path() -> Path:
    """
    1) Si EXCEL_TEMPLATE_PATH está en .env y existe, usar eso.
    2) Si hay exactamente un .xlsx en data/templates, usarlo.
    3) Si no, usar DEFAULT_TEMPLATE_PATH.
    """
    env_path = os.getenv("EXCEL_TEMPLATE_PATH")
    if env_path:
        p = Path(env_path)
        if not p.is_absolute():
            p = ROOT_BACKEND / p
        if p.exists():
            return p
    if TEMPLATES_DIR.exists():
        xlsx = list(TEMPLATES_DIR.glob("*.xlsx"))
        if len(xlsx) == 1:
            return xlsx[0]
    return DEFAULT_TEMPLATE_PATH

# ---------------- Limpieza estética / placeholders ----------------
HEADER_TOKENS = {
    "objetivo de la prueba", "objetivo",
    "funcionalidad",
    "resultado esperado", "resultado",
    "validación del resultado", "validacion del resultado",
    "observaciones",
    "caso de prueba", "id. caso de prueba"
}
FILLER_RE = re.compile(r"""
    ^\s*[-_\.·]{3,}\s*$                            # -----, _____, ..., ···
    |^\s*(resultado\s+esperado
         |objetivo\s+de\s+la\s+prueba
         |funcionalidad
         |observaciones
         |validaci[oó]n\s+del\s+resultado
         |caso\s+de\s+prueba
         |id\.\s*caso\s*de\s*prueba)\s*$          # eco de headers
""", re.IGNORECASE | re.VERBOSE)

def _strip_header_echo(value: str) -> str:
    """Limpia celdas que quedaron con encabezados/placeholder o rayas."""
    v = _sanitize(value)
    if not v:
        return ""
    if v.lower() in HEADER_TOKENS:
        return ""
    if FILLER_RE.match(v):
        return ""
    v = re.sub(r"[-_\.·]{3,}", " ", v).strip()  # rayas y rellenos
    v = re.sub(r"\s{2,}", " ", v)               # espacios múltiples
    return v

# --------------- Localización de headers ----------------
@dataclass
class HeaderLocator:
    sheet: Worksheet
    header_row: int
    col_by_header: Dict[str, int]  # normalized header -> col (1-based)

def _find_headers_anywhere(ws: Worksheet, target_headers: List[str]) -> Optional[HeaderLocator]:
    target_norm = [_norm(h) for h in target_headers]
    best: Optional[Tuple[int, Dict[str,int]]] = None
    max_row = ws.max_row or 200
    max_col = ws.max_column or 30

    for r in range(1, max_row + 1):
        row_map: Dict[str, int] = {}
        matches = 0
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                nv = _norm(v)
                for th in target_norm:
                    if nv == th and th not in row_map:
                        row_map[th] = c
                        matches += 1
        if matches > 0 and (best is None or matches > len(best[1])):
            best = (r, row_map)

    if best and len(best[1]) >= 3:
        col_by_header: Dict[str, int] = {}
        for raw in target_headers:
            nraw = _norm(raw)
            if nraw in best[1]:
                col_by_header[_norm(raw)] = best[1][nraw]
        return HeaderLocator(ws, best[0], col_by_header)
    return None

def _pick_sheet_and_headers(wb: Workbook) -> HeaderLocator:
    desired = os.getenv("EXCEL_SHEET_NAME")
    if desired and desired in wb.sheetnames:
        ws = wb[desired]
        loc = _find_headers_anywhere(ws, TEMPLATE_HEADERS)
        if loc:
            return loc

    candidates = []
    for name in wb.sheetnames:
        ws = wb[name]
        loc = _find_headers_anywhere(ws, TEMPLATE_HEADERS)
        if loc:
            score = len(loc.col_by_header) - (1 if "bugs" in name.lower() else 0)
            candidates.append((score, loc))
    if not candidates:
        raise RuntimeError("No pude localizar los encabezados del template en ninguna hoja.")
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]

# --------------- Funcionalidad: inferencia ----------------
def _infer_funcionalidad(caso: str, objetivo: str, esperado: str) -> str:
    if EXCEL_FUNC_OVERRIDE:
        return EXCEL_FUNC_OVERRIDE

    base_caso = _strip_header_echo(caso)
    if base_caso and base_caso not in ("-", "—"):
        s = re.sub(r"^(validar|verificar|comprobar|chequear)\s+(que\s+)?", "", base_caso, flags=re.IGNORECASE)
        s = re.sub(r"\bcp[-_ ]?\d+\b$", "", s, flags=re.IGNORECASE).strip(" -—:.,")
        if len(s) >= 4:
            return _title_case_es(s)

    text = " ".join([caso or "", objetivo or "", esperado or ""])
    t = _sanitize(text)

    m_code = re.search(r"\b([A-Za-z]{3,}\d{2,})\b", t)
    code = m_code.group(1).upper() if m_code else ""

    plan_nums = re.findall(r"\bplanilla[s]?\s*(\d{3})\b", t, flags=re.IGNORECASE)
    plan_str = ""
    if plan_nums:
        seen = []
        for n in plan_nums:
            if n not in seen:
                seen.append(n)
        plan_str = ("planilla " + " y ".join(seen[:3])) if seen else ""

    carga = "carga masiva" if re.search(r"\bcarga\s+masiva\b", t, flags=re.IGNORECASE) else ""

    parts = []
    if code: parts.append(code)
    desc = " ".join(x for x in [carga, plan_str] if x).strip()
    if desc: parts.append(_title_case_es(desc))

    if parts:
        return " – ".join(parts)

    if EXCEL_FUNC_DEFAULT:
        return EXCEL_FUNC_DEFAULT
    return "—"

# --------------- Formateadores por columna ----------------
def _format_objetivo(objetivo: str, funcionalidad: str, esperado: str) -> str:
    obj = _strip_header_echo(objetivo)
    if not obj:
        base = _strip_header_echo(funcionalidad) or "la funcionalidad"
        obj = f"Validar {base.lower()}"
    obj = re.sub(r"^(verificar|comprobar|chequear)\b", "Validar", obj, flags=re.IGNORECASE)
    return obj.rstrip(" .")

def _format_resultado(esperado: str) -> str:
    t = _strip_header_echo(esperado)
    if not t:
        return ""   # evitar genéricos
    return _title_case_es(t)

def _classify_observaciones(esperado: str, objetivo: str) -> str:
    if not OBS_CLASSIFY:
        return ""
    e = _sanitize(esperado).lower()
    o = _sanitize(objetivo)
    neg = any(k in e for k in ["rechaz", "error", "inválid", "invalido", "incorrect"])
    pos = any(k in e for k in ["proces", "correct", "exitos", "acept"])
    hint = ""
    m = re.search(r"\b([A-Z]{3,}[A-Z0-9_-]*\s*\d+)\b", o)
    if m:
        hint = m.group(1).strip()
    if neg and not pos:
        return f"Negativo{(' – ' + hint) if hint else ''}"
    if pos and not neg:
        return f"Positivo{(' – ' + hint) if hint else ''}"
    return ""

def _format_observaciones(prioridad: str, esperado: str, objetivo: str) -> str:
    base_text = _strip_header_echo(esperado)
    base = _classify_observaciones(base_text, objetivo) if base_text else ""
    pr = _sanitize(prioridad)
    parts = []
    if base: parts.append(base)
    if pr: parts.append(f"Prioridad: {pr}")
    return " – ".join(parts) if parts else ""

# --------------- Utilidades Excel ----------------
PLACEHOLDER_HASH_RE = re.compile(r"^\s*#")
def _top_left_of_merge(ws: Worksheet, r: int, c: int) -> Tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return rng.min_row, rng.min_col
    return r, c

# --------------- Mapeo fila -> template ----------------
def _compose_template_row(src: Dict[str, Any], idx: int) -> Dict[str, Any]:
    src_func = (src.get("Funcionalidad") or "").strip()

    if EXCEL_FUNC_OVERRIDE:
        funcionalidad = EXCEL_FUNC_OVERRIDE
    elif EXCEL_FUNC_FORCE:
        funcionalidad = src_func or _infer_funcionalidad(
            src.get("Caso de Prueba", ""),
            src.get("Objetivo", ""),
            src.get("Resultado Esperado", ""),
        ) or (EXCEL_FUNC_DEFAULT or "—")
    else:
        funcionalidad = src_func or _infer_funcionalidad(
            src.get("Caso de Prueba", ""),
            src.get("Objetivo", ""),
            src.get("Resultado Esperado", ""),
        ) or (EXCEL_FUNC_DEFAULT or "—")

    objetivo = _format_objetivo(src.get("Objetivo", ""), funcionalidad, src.get("Resultado Esperado", ""))
    esperado = _format_resultado(src.get("Resultado Esperado", ""))
    observ   = _format_observaciones(src.get("Prioridad", ""), src.get("Resultado Esperado", ""), src.get("Objetivo", ""))

    return {
        _norm("Id. Caso de Prueba"): f"CP-{idx:03d}" if ID_STYLE == "cp" else idx,
        _norm("Objetivo de la prueba"): objetivo,
        _norm("Funcionalidad"): funcionalidad,
        _norm("Resultado Esperado"): esperado,
        _norm("Validación del resultado"): "" if LEAVE_VALIDACION_EMPTY else "",
        _norm("Observaciones"): observ,
    }

# --------------- Escritura en el template ----------------
def _write_with_template(rows: List[Dict[str, Any]], template_path: Path) -> bytes:
    wb = load_workbook(str(template_path))
    loc = _pick_sheet_and_headers(wb)
    ws = loc.sheet

    # Arranca debajo del header
    r = loc.header_row + 1 if FILL_FROM_TEMPLATE_BOXES else loc.header_row + 1

    id_norm    = _norm("Id. Caso de Prueba")
    valid_norm = _norm("Validación del resultado")
    func_norm  = _norm("Funcionalidad")

    func_col = loc.col_by_header.get(func_norm, None)
    if func_col is None and EXCEL_FUNC_COL > 0:
        func_col = EXCEL_FUNC_COL

    for i, src in enumerate(rows, start=1):
        mapped = _compose_template_row(src, idx=i)

        for th_norm, col in loc.col_by_header.items():
            value = mapped.get(th_norm, None)
            if value is None:
                continue

            # ID: si ya hay número, no pisar
            if th_norm == id_norm:
                existing = ws.cell(r, col).value
                if isinstance(existing, (int, float)) and str(existing).strip():
                    continue
                if existing not in (None, ""):
                    continue

            # Validación: vacía
            if th_norm == valid_norm and LEAVE_VALIDACION_EMPTY:
                rr, cc = _top_left_of_merge(ws, r, col)
                ws.cell(rr, cc).value = ""
                continue

            # Funcionalidad
            if th_norm == func_norm:
                rr, cc = _top_left_of_merge(ws, r, col)
                val = _strip_header_echo(value) if isinstance(value, str) else value
                if EXCEL_FUNC_FORCE:
                    ws.cell(rr, cc).value = val or EXCEL_FUNC_DEFAULT or "—"
                else:
                    existing = ws.cell(rr, cc).value
                    if not existing or (isinstance(existing, str) and PLACEHOLDER_HASH_RE.search(existing)):
                        ws.cell(rr, cc).value = val
                continue

            # Resto normal
            rr, cc = _top_left_of_merge(ws, r, col)
            clean = _strip_header_echo(value) if isinstance(value, str) else value
            ws.cell(rr, cc).value = clean

        # Si el header 'Funcionalidad' no existe, escribir en la col fallback
        if func_col and (func_norm not in loc.col_by_header):
            v = mapped.get(func_norm, EXCEL_FUNC_DEFAULT or "—")
            rr, cc = _top_left_of_merge(ws, r, func_col)
            existing = ws.cell(rr, cc).value
            if EXCEL_FUNC_FORCE or not existing or (isinstance(existing, str) and PLACEHOLDER_HASH_RE.search(existing)):
                ws.cell(rr, cc).value = _strip_header_echo(v)

        # Wrap + alineación arriba para columnas largas
        for key in (_norm("Resultado Esperado"), _norm("Observaciones")):
            if key in loc.col_by_header:
                cidx = loc.col_by_header[key]
                rr, cc = _top_left_of_merge(ws, r, cidx)
                cell = ws.cell(rr, cc)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        r += 1

    # Ajuste de anchos (estético)
    for header_norm, col in loc.col_by_header.items():
        width = 18
        if header_norm in {_norm("Objetivo de la prueba"), _norm("Resultado Esperado"), _norm("Observaciones")}:
            width = 36
        ws.column_dimensions[get_column_letter(col)].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# --------------- Fallback simple ----------------
def _write_simple(rows: List[Dict[str, Any]]) -> bytes:
    df = pd.DataFrame(rows)
    for col in SOURCE_HEADERS:
        if col not in df.columns:
            df[col] = ""
    df = df[SOURCE_HEADERS]
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Casos generados")
    bio.seek(0)
    return bio.getvalue()

# ---------------- API pública ----------------
def to_excel_bytes(rows: List[Dict[str, Any]]) -> bytes:
    path = resolve_template_path()
    try:
        if path.exists():
            print(f"[excel] Usando plantilla: {path}")
            return _write_with_template(rows, path)
        print(f"[excel] Plantilla NO encontrada en: {path}. Usando fallback simple.")
        return _write_simple(rows)
    except Exception as e:
        print(f"[excel] Error usando plantilla ({path}): {e}. Usando fallback simple.")
        return _write_simple(rows)
