from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from pymongo import ASCENDING, DESCENDING

from services.auth import _db, current_user


router = APIRouter(prefix="/api/quality-records", tags=["quality-records"])

ALLOWED_ROLES = {"qa", "lider"}
COLLECTION = "QualityQaRecords"

COLUMNS = [
    ("id_req", "ID REQ"),
    ("requirement_name", "Nombre Requerimiento"),
    ("qa_responsible", "Responsable QA"),
    ("design_time", "Tiempo de Diseno"),
    ("generated_cases", "Casos Generados"),
    ("ok_cases", "Casos OK"),
    ("ai_quality_percent", "% de Calidad IA"),
    ("additional_qa_cases", "Casos Adicionales QA"),
    ("post_qa_review_quality_percent", "% de Calidad Post Revision QA"),
    ("additional_functional_cases", "Casos Adicionales Funcional"),
    ("post_functional_review_quality_percent", "% de Calidad Post Revision Funcional"),
]


class QualityRecordIn(BaseModel):
    id_req: str
    requirement_name: str
    design_time: Optional[float] = None
    design_time_seconds: Optional[int] = None
    generated_cases: int
    ok_cases: int
    additional_qa_cases: int
    additional_functional_cases: int = 0
    ai_quality_percent: Optional[float] = None
    post_qa_review_quality_percent: Optional[float] = None
    post_functional_review_quality_percent: Optional[float] = None


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _require_allowed(user: Dict[str, Any]) -> None:
    if user.get("role") not in ALLOWED_ROLES:
        raise HTTPException(status_code=403, detail="Registro IA disponible solo para QA o Lider.")


def _oid(record_id: str) -> ObjectId:
    if not ObjectId.is_valid(record_id):
        raise HTTPException(status_code=400, detail="ID invalido.")
    return ObjectId(record_id)


def _clean_text(value: str, field: str, max_len: int = 220) -> str:
    clean = (value or "").replace("\x00", "").strip()
    if not clean:
        raise HTTPException(status_code=400, detail=f"{field} es obligatorio.")
    return clean[:max_len]


def _non_negative_number(value: float, field: str) -> float:
    if value is None or value < 0:
        raise HTTPException(status_code=400, detail=f"{field} no puede ser negativo.")
    return value


def _non_negative_int(value: int, field: str) -> int:
    if value is None or int(value) < 0:
        raise HTTPException(status_code=400, detail=f"{field} no puede ser negativo.")
    return int(value)


def _pct(value: float) -> float:
    return round(value, 2)


def _calculated(
    generated_cases: int,
    ok_cases: int,
    additional_qa_cases: int,
    additional_functional_cases: int,
) -> Dict[str, float]:
    if generated_cases <= 0:
        raise HTTPException(status_code=400, detail="Casos Generados debe ser mayor a 0.")
    if ok_cases > generated_cases:
        raise HTTPException(status_code=400, detail="Casos OK no puede ser mayor a Casos Generados.")

    qa_total = ok_cases + additional_qa_cases
    if qa_total <= 0:
        raise HTTPException(status_code=400, detail="Casos OK + Casos Adicionales QA debe ser mayor a 0.")

    functional_total = qa_total + additional_functional_cases
    if functional_total <= 0:
        raise HTTPException(status_code=400, detail="El total post revision funcional debe ser mayor a 0.")

    return {
        "ai_quality_percent": _pct((ok_cases / generated_cases) * 100),
        "post_qa_review_quality_percent": _pct((ok_cases / qa_total) * 100),
        "post_functional_review_quality_percent": _pct((ok_cases / functional_total) * 100),
    }


def _payload(payload: QualityRecordIn, user: Dict[str, Any]) -> Dict[str, Any]:
    generated = _non_negative_int(payload.generated_cases, "Casos Generados")
    ok = _non_negative_int(payload.ok_cases, "Casos OK")
    additional_qa = _non_negative_int(payload.additional_qa_cases, "Casos Adicionales QA")
    additional_functional = _non_negative_int(
        payload.additional_functional_cases,
        "Casos Adicionales Funcional",
    )
    if payload.design_time_seconds is not None:
        design_time_seconds = _non_negative_int(payload.design_time_seconds, "Tiempo de Diseno")
    elif payload.design_time is not None:
        # Compatibilidad con registros anteriores, donde el valor representaba minutos.
        design_time_seconds = round(_non_negative_number(float(payload.design_time), "Tiempo de Diseno") * 60)
    else:
        raise HTTPException(status_code=400, detail="Tiempo de Diseno es obligatorio.")

    data = {
        "id_req": _clean_text(payload.id_req, "ID REQ"),
        "requirement_name": _clean_text(payload.requirement_name, "Nombre Requerimiento"),
        "qa_responsible": user["username"],
        "design_time_seconds": design_time_seconds,
        "design_time": round(design_time_seconds / 60, 2),
        "generated_cases": generated,
        "ok_cases": ok,
        "additional_qa_cases": additional_qa,
        "additional_functional_cases": additional_functional,
    }
    data.update(_calculated(generated, ok, additional_qa, additional_functional))
    return data


def _public(doc: Dict[str, Any]) -> Dict[str, Any]:
    doc = dict(doc)
    if "design_time_seconds" not in doc:
        doc["design_time_seconds"] = round(float(doc.get("design_time") or 0) * 60)
    if "post_qa_review_quality_percent" not in doc and "post_review_quality_percent" in doc:
        doc["post_qa_review_quality_percent"] = doc.get("post_review_quality_percent")
    if "additional_functional_cases" not in doc:
        doc["additional_functional_cases"] = 0
    if "post_functional_review_quality_percent" not in doc:
        generated = int(doc.get("generated_cases") or 0)
        ok = int(doc.get("ok_cases") or 0)
        additional_qa = int(doc.get("additional_qa_cases") or 0)
        additional_functional = int(doc.get("additional_functional_cases") or 0)
        try:
            doc.update(_calculated(generated, ok, additional_qa, additional_functional))
        except HTTPException:
            doc["post_functional_review_quality_percent"] = 0
    doc["id"] = str(doc.pop("_id"))
    return doc


async def _ensure_indexes() -> None:
    db = _db()
    await db[COLLECTION].create_index([("qa_responsible", ASCENDING), ("created_at", DESCENDING)])
    await db[COLLECTION].create_index([("id_req", ASCENDING)])


@router.get("")
async def list_records(user: Dict[str, Any] = Depends(current_user)):
    _require_allowed(user)
    await _ensure_indexes()
    docs = await _db()[COLLECTION].find({}).sort("created_at", DESCENDING).to_list(1000)
    return {"records": [_public(doc) for doc in docs]}


@router.post("")
async def create_record(payload: QualityRecordIn, user: Dict[str, Any] = Depends(current_user)):
    _require_allowed(user)
    if user["role"] != "qa":
        raise HTTPException(status_code=403, detail="Solo QA puede crear registros.")
    await _ensure_indexes()
    now = _now()
    doc = {
        **_payload(payload, user),
        "created_by": user["username"],
        "created_at": now,
        "updated_at": now,
    }
    result = await _db()[COLLECTION].insert_one(doc)
    doc["_id"] = result.inserted_id
    return {"record": _public(doc)}


@router.put("/{record_id}")
async def update_record(record_id: str, payload: QualityRecordIn, user: Dict[str, Any] = Depends(current_user)):
    _require_allowed(user)
    if user["role"] != "qa":
        raise HTTPException(status_code=403, detail="Solo QA puede editar registros.")
    db = _db()
    current = await db[COLLECTION].find_one({"_id": _oid(record_id)})
    if not current:
        raise HTTPException(status_code=404, detail="Registro no encontrado.")
    if current.get("created_by") != user["username"]:
        raise HTTPException(status_code=403, detail="Solo podes editar registros creados por vos.")
    update = {**_payload(payload, user), "updated_at": _now()}
    await db[COLLECTION].update_one({"_id": current["_id"]}, {"$set": update})
    current.update(update)
    return {"record": _public(current)}


@router.delete("/{record_id}")
async def delete_record(record_id: str, user: Dict[str, Any] = Depends(current_user)):
    _require_allowed(user)
    if user["role"] != "qa":
        raise HTTPException(status_code=403, detail="Solo QA puede eliminar registros.")
    db = _db()
    current = await db[COLLECTION].find_one({"_id": _oid(record_id)})
    if not current:
        raise HTTPException(status_code=404, detail="Registro no encontrado.")
    if current.get("created_by") != user["username"]:
        raise HTTPException(status_code=403, detail="Solo podes eliminar registros creados por vos.")
    await db[COLLECTION].delete_one({"_id": current["_id"]})
    return {"ok": True}


@router.get("/export")
async def export_records(user: Dict[str, Any] = Depends(current_user)):
    _require_allowed(user)
    await _ensure_indexes()
    docs = await _db()[COLLECTION].find({}).sort("created_at", DESCENDING).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro IA"
    ws.append([label for _, label in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")

    for doc in docs:
        public_doc = _public(doc)
        values = []
        for key, _ in COLUMNS:
            if key == "design_time":
                values.append(public_doc.get("design_time_seconds", 0) / 86400)
            else:
                values.append(public_doc.get(key, ""))
        ws.append(values)

    percent_cols = [7, 9, 11]
    for row in ws.iter_rows(min_row=2):
        row[3].number_format = '[m]" min "ss" seg"'
        for col in percent_cols:
            row[col - 1].number_format = '0.00"%"'

    widths = [16, 32, 22, 18, 18, 14, 18, 24, 28, 28, 34]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="registro_ia.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
