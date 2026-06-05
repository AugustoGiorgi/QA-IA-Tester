from __future__ import annotations

import csv
import io
import json
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from uuid import uuid4

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook
from pymongo import ASCENDING, UpdateOne
from pymongo.errors import BulkWriteError

from services.files import safe_filename


load_dotenv()

router = APIRouter(prefix="/api/import-data", tags=["data-import"])

BACKEND_DIR = Path(__file__).resolve().parent.parent
OUTPUTS_DIR = BACKEND_DIR / "data" / "outputs"
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

MAX_IMPORT_FILE_MB = int(os.getenv("MAX_IMPORT_FILE_MB", "20"))
MAX_IMPORT_BYTES = MAX_IMPORT_FILE_MB * 1024 * 1024
DEFAULT_BATCH_SIZE = int(os.getenv("IMPORT_BATCH_SIZE", "500"))
MONGO_URI = os.getenv("MONGO_URI")
MONGO_DB = os.getenv("MONGO_DB", "Life_Projects")

COLLECTION_RE = re.compile(r"^[A-Za-z0-9_\-]{1,80}$")
FIELD_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_ .\-]{0,120}$")
SUPPORTED_TYPES = {"string", "int", "float", "bool", "date"}

_client: Optional[AsyncIOMotorClient] = None


@dataclass
class RowError:
    row: int
    field: str
    error: str
    value: str = ""


@dataclass
class ParsedRow:
    row_number: int
    data: Dict[str, Any]


@dataclass
class ImportConfig:
    collection: str
    unique_key: str
    required_fields: List[str] = field(default_factory=list)
    field_types: Dict[str, str] = field(default_factory=dict)
    batch_size: int = DEFAULT_BATCH_SIZE


@dataclass
class ImportResult:
    total_rows: int = 0
    valid_rows: int = 0
    empty_rows: int = 0
    created: int = 0
    updated: int = 0
    errors: List[RowError] = field(default_factory=list)
    error_report: Optional[str] = None


def _mongo_db():
    global _client
    if not MONGO_URI:
        raise HTTPException(status_code=500, detail="MONGO_URI no está configurado en el backend.")
    if _client is None:
        _client = AsyncIOMotorClient(MONGO_URI)
    return _client[MONGO_DB]


def _split_fields(raw: str) -> List[str]:
    return [item.strip() for item in (raw or "").split(",") if item.strip()]


def _load_field_types(raw: str) -> Dict[str, str]:
    if not raw or not raw.strip():
        return {}
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        data = {}
        for part in _split_fields(raw):
            if ":" not in part:
                raise HTTPException(status_code=400, detail=f"Tipo inválido: {part}")
            key, value = part.split(":", 1)
            data[key.strip()] = value.strip()

    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="field_types debe ser un objeto JSON o pares campo:tipo.")

    normalized: Dict[str, str] = {}
    for key, value in data.items():
        field = _sanitize_field_name(str(key))
        typ = str(value).strip().lower()
        if typ not in SUPPORTED_TYPES:
            raise HTTPException(status_code=400, detail=f"Tipo no soportado para {field}: {typ}")
        normalized[field] = typ
    return normalized


def _sanitize_field_name(name: str) -> str:
    clean = re.sub(r"[\x00-\x1f]", "", str(name or "")).strip()
    clean = clean.replace("$", "").replace(".", "_")
    clean = re.sub(r"\s+", " ", clean)
    if not clean or not FIELD_RE.fullmatch(clean):
        raise HTTPException(status_code=400, detail=f"Nombre de campo inválido: {name}")
    return clean


def _sanitize_collection(name: str) -> str:
    clean = (name or "").strip()
    if not COLLECTION_RE.fullmatch(clean):
        raise HTTPException(status_code=400, detail="Nombre de colección inválido.")
    return clean


def _sanitize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        clean = value.replace("\x00", "").strip()
        return clean if clean != "" else None
    return value


def _is_empty_row(row: Dict[str, Any]) -> bool:
    return all(_sanitize_value(value) is None for value in row.values())


def _parse_config(
    collection: str,
    unique_key: str,
    required_fields: str,
    field_types: str,
) -> ImportConfig:
    normalized_unique = _sanitize_field_name(unique_key)
    required = [_sanitize_field_name(field) for field in _split_fields(required_fields)]
    if normalized_unique not in required:
        required.insert(0, normalized_unique)
    return ImportConfig(
        collection=_sanitize_collection(collection),
        unique_key=normalized_unique,
        required_fields=required,
        field_types=_load_field_types(field_types),
    )


def _coerce_value(value: Any, typ: str) -> Tuple[Any, Optional[str]]:
    value = _sanitize_value(value)
    if value is None:
        return None, None

    try:
        if typ == "string":
            return str(value), None
        if typ == "int":
            if isinstance(value, float) and value.is_integer():
                return int(value), None
            return int(str(value).replace(",", "").strip()), None
        if typ == "float":
            return float(str(value).replace(",", ".").strip()), None
        if typ == "bool":
            normalized = str(value).strip().lower()
            if normalized in {"1", "true", "si", "sí", "yes", "y"}:
                return True, None
            if normalized in {"0", "false", "no", "n"}:
                return False, None
            return None, "debe ser booleano"
        if typ == "date":
            if isinstance(value, datetime):
                return value, None
            raw = str(value).strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(raw, fmt), None
                except ValueError:
                    pass
            return None, "debe ser fecha válida"
    except (TypeError, ValueError):
        return None, f"debe ser {typ}"

    return value, None


def _normalize_row(raw: Dict[str, Any], row_number: int, config: ImportConfig) -> Tuple[Optional[ParsedRow], List[RowError]]:
    errors: List[RowError] = []
    row: Dict[str, Any] = {}

    for key, value in raw.items():
        if key is None:
            continue
        field_name = _sanitize_field_name(str(key))
        clean_value = _sanitize_value(value)
        if clean_value is not None:
            row[field_name] = clean_value

    if _is_empty_row(row):
        return None, []

    for field in config.required_fields:
        if row.get(field) in (None, ""):
            errors.append(RowError(row=row_number, field=field, error="campo requerido faltante"))

    for field, typ in config.field_types.items():
        if field not in row:
            continue
        coerced, error = _coerce_value(row[field], typ)
        if error:
            errors.append(RowError(row=row_number, field=field, error=error, value=str(row[field])))
        elif coerced is not None:
            row[field] = coerced

    if errors:
        return None, errors

    return ParsedRow(row_number=row_number, data=row), []


def _parse_csv(contents: bytes) -> Iterable[ParsedRow]:
    try:
        text = contents.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = contents.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="El CSV no tiene encabezados.")

    for index, row in enumerate(reader, start=2):
        yield ParsedRow(row_number=index, data=dict(row))


def _parse_xlsx(contents: bytes) -> Iterable[ParsedRow]:
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    header_row_number = 0
    headers: List[str] = []
    for row_number, values in enumerate(rows, start=1):
        if values and any(_sanitize_value(value) is not None for value in values):
            headers = [_sanitize_field_name(str(value)) if value is not None else "" for value in values]
            header_row_number = row_number
            break

    if not headers:
        raise HTTPException(status_code=400, detail="El Excel no tiene encabezados.")

    for row_number, values in enumerate(rows, start=header_row_number + 1):
        row = {
            headers[index]: values[index] if index < len(values) else None
            for index in range(len(headers))
            if headers[index]
        }
        yield ParsedRow(row_number=row_number, data=row)


def _parse_file(filename: str, contents: bytes) -> Iterable[ParsedRow]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _parse_csv(contents)
    if lower.endswith(".xlsx"):
        return _parse_xlsx(contents)
    raise HTTPException(status_code=400, detail="Formato inválido. Subí un archivo .csv o .xlsx.")


def _write_error_report(errors: List[RowError]) -> Optional[str]:
    if not errors:
        return None

    report_id = uuid4().hex
    filename = f"import_errors_{report_id}.csv"
    path = OUTPUTS_DIR / filename
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["row", "field", "error", "value"])
        writer.writeheader()
        for err in errors:
            writer.writerow({
                "row": err.row,
                "field": err.field,
                "error": err.error,
                "value": err.value,
            })
    return filename


async def _flush_batch(collection, operations: List[UpdateOne], row_numbers: List[int], result: ImportResult) -> None:
    if not operations:
        return
    try:
        bulk_result = await collection.bulk_write(operations, ordered=False)
        result.created += bulk_result.upserted_count
        result.updated += bulk_result.matched_count
    except BulkWriteError as exc:
        details = exc.details or {}
        for item in details.get("writeErrors", []):
            idx = int(item.get("index", 0))
            row = row_numbers[idx] if idx < len(row_numbers) else 0
            result.errors.append(RowError(
                row=row,
                field="mongodb",
                error=item.get("errmsg", "error de escritura"),
            ))
        partial = details.get("nInserted", 0)
        result.created += int(partial or 0)


async def import_file(contents: bytes, filename: str, config: ImportConfig) -> ImportResult:
    db = _mongo_db()
    collection = db[config.collection]
    await collection.create_index([(config.unique_key, ASCENDING)], unique=True, name=f"uniq_{config.unique_key}")

    result = ImportResult()
    seen_keys = set()
    operations: List[UpdateOne] = []
    row_numbers: List[int] = []

    for raw_row in _parse_file(filename, contents):
        result.total_rows += 1
        parsed, errors = _normalize_row(raw_row.data, raw_row.row_number, config)
        if not parsed:
            if errors:
                result.errors.extend(errors)
            else:
                result.empty_rows += 1
            continue

        key_value = parsed.data.get(config.unique_key)
        if key_value in seen_keys:
            result.errors.append(RowError(
                row=parsed.row_number,
                field=config.unique_key,
                error="clave duplicada dentro del archivo",
                value=str(key_value),
            ))
            continue
        seen_keys.add(key_value)

        update_doc = {k: v for k, v in parsed.data.items() if v is not None}
        operations.append(UpdateOne(
            {config.unique_key: key_value},
            {"$set": update_doc, "$setOnInsert": {"created_at": datetime.utcnow()}},
            upsert=True,
        ))
        row_numbers.append(parsed.row_number)
        result.valid_rows += 1

        if len(operations) >= config.batch_size:
            await _flush_batch(collection, operations, row_numbers, result)
            operations, row_numbers = [], []

    await _flush_batch(collection, operations, row_numbers, result)
    result.error_report = _write_error_report(result.errors)
    return result


@router.post("")
async def import_data(
    file: UploadFile = File(...),
    collection: str = Form(...),
    unique_key: str = Form(...),
    required_fields: str = Form(""),
    field_types: str = Form(""),
):
    filename = safe_filename(file.filename)
    contents = await file.read()
    if len(contents) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"El archivo supera el máximo permitido de {MAX_IMPORT_FILE_MB} MB.",
        )

    config = _parse_config(collection, unique_key, required_fields, field_types)
    result = await import_file(contents, filename, config)
    return {
        "filename": filename,
        "collection": config.collection,
        "unique_key": config.unique_key,
        "total_rows": result.total_rows,
        "valid_rows": result.valid_rows,
        "empty_rows": result.empty_rows,
        "created": result.created,
        "updated": result.updated,
        "error_count": len(result.errors),
        "errors": [err.__dict__ for err in result.errors[:100]],
        "error_report": result.error_report,
    }
