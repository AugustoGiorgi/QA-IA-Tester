from __future__ import annotations

import json
import re
import shlex
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import parse_qsl, urlparse

import yaml
from bson import ObjectId
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import JSONResponse, PlainTextResponse, Response, StreamingResponse
from pydantic import BaseModel
from pymongo import DESCENDING

from services.activity import record_activity
from services.auth import _db, current_user
from services.files import safe_filename
from services.parsing import docx_to_text


router = APIRouter(prefix="/api/postman", tags=["postman-generator"])
COLLECTION = "PostmanGenerationDrafts"
OUTPUT_NAME = "qa_postman_package"
POSTMAN_SCHEMA = "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
HTTP_METHODS = {"GET", "POST", "PUT", "PATCH", "DELETE", "HEAD", "OPTIONS"}
TEXT_EXTENSIONS = {".txt", ".md", ".json", ".yaml", ".yml", ".xml", ".csv", ".curl", ".http"}
SECRET_RE = re.compile(
    r"(?i)(bearer\s+[a-z0-9._\-]{12,}|api[_-]?key\s*[:=]\s*[^\s,;]{8,}|"
    r"password\s*[:=]\s*[^\s,;]{4,}|secret\s*[:=]\s*[^\s,;]{6,}|"
    r"token\s*[:=]\s*[^\s,;]{8,}|-----BEGIN\s+(?:RSA\s+)?PRIVATE\s+KEY-----)"
)


class DraftUpdate(BaseModel):
    endpoints: Optional[List[Dict[str, Any]]] = None
    test_cases: Optional[List[Dict[str, Any]]] = None
    associations: Optional[List[Dict[str, Any]]] = None
    variables: Optional[List[Dict[str, Any]]] = None
    warnings: Optional[List[Dict[str, Any]]] = None
    conflicts: Optional[List[Dict[str, Any]]] = None
    folders: Optional[List[str]] = None


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _oid(value: str) -> ObjectId:
    if not ObjectId.is_valid(value):
        raise HTTPException(status_code=400, detail="ID invalido.")
    return ObjectId(value)


def _clean(value: Any, max_len: int = 12000) -> str:
    return str(value or "").replace("\x00", "").strip()[:max_len]


def _safe_key(value: str, fallback: str = "variable") -> str:
    parts = re.findall(r"[A-Za-z0-9]+", value or "")
    if not parts:
        return fallback
    first, *rest = parts
    return first[:1].lower() + first[1:] + "".join(part[:1].upper() + part[1:] for part in rest)


def _source_ref(name: str, kind: str, location: str = "") -> Dict[str, str]:
    return {"source": name, "kind": kind, "location": location}


async def _read_upload(file: UploadFile) -> Dict[str, Any]:
    filename = safe_filename(file.filename or "fuente.txt")
    suffix = Path(filename).suffix.lower()
    raw = await file.read()
    text = ""
    parse_warning = ""
    try:
        if suffix == ".docx":
            tmp = Path(__file__).resolve().parent.parent / "data" / f"postman_tmp_{datetime.utcnow().timestamp()}_{filename}"
            tmp.write_bytes(raw)
            try:
                text = docx_to_text(str(tmp))
            finally:
                tmp.unlink(missing_ok=True)
        elif suffix == ".xlsx":
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
                lines: List[str] = []
                for sheet in workbook.worksheets:
                    lines.append(f"HOJA: {sheet.title}")
                    for row in sheet.iter_rows(values_only=True):
                        values = [str(value).strip() for value in row if value not in (None, "")]
                        if values:
                            lines.append(" | ".join(values))
                text = "\n".join(lines)
            except Exception:
                parse_warning = "No se pudo leer el Excel. Verifica que sea .xlsx valido."
        elif suffix == ".pdf":
            try:
                from pypdf import PdfReader

                reader = PdfReader(BytesIO(raw))
                text = "\n".join(page.extract_text() or "" for page in reader.pages)
            except Exception:
                parse_warning = "No se pudo extraer texto del PDF. Si es escaneado, converti a texto o DOCX."
        elif suffix in TEXT_EXTENSIONS or not suffix:
            text = raw.decode("utf-8", errors="replace")
        else:
            parse_warning = f"Formato {suffix} recibido; se conservara como fuente pero no se pudo leer como texto."
    except Exception as exc:
        parse_warning = f"No se pudo leer {filename}: {type(exc).__name__}."
    return {
        "name": filename,
        "extension": suffix,
        "text": _clean(text, 300000),
        "size": len(raw),
        "warning": parse_warning,
    }


def _try_load_structured(text: str) -> Optional[Any]:
    clean = (text or "").strip()
    if not clean:
        return None
    for loader in (json.loads, yaml.safe_load):
        try:
            data = loader(clean)
        except Exception:
            continue
        if isinstance(data, (dict, list)):
            return data
    return None


def _classify_source(source: Dict[str, Any]) -> List[str]:
    text = source.get("text", "")
    normalized = text.lower()
    data = _try_load_structured(text)
    kinds: List[str] = []
    if isinstance(data, dict) and ("openapi" in data or "swagger" in data or "paths" in data):
        kinds.append("api_definition")
    if isinstance(data, dict) and data.get("info", {}).get("schema", "").endswith("collection/v2.1.0/collection.json"):
        kinds.append("postman_collection")
    if "curl " in normalized:
        kinds.append("curl")
    if re.search(r"\b(get|post|put|patch|delete|head|options)\s+https?://", normalized):
        kinds.append("api_examples")
    if re.search(r"\b(caso|test case|prueba|expected|resultado esperado|precondicion)", normalized):
        kinds.append("test_cases")
    if re.search(r"\{\{[a-zA-Z0-9_.-]+\}\}|environment|base url|variable", normalized):
        kinds.append("variables")
    if SECRET_RE.search(text):
        kinds.append("sensitive_data")
    if not kinds:
        kinds.append("complementary_documentation")
    return list(dict.fromkeys(kinds))


def _infer_content_type(headers: List[Dict[str, str]], body: Any) -> str:
    for header in headers:
        if header.get("key", "").lower() == "content-type":
            return header.get("value", "")
    if isinstance(body, (dict, list)):
        return "application/json"
    if isinstance(body, str) and body.strip().startswith("<"):
        return "application/xml"
    if body not in (None, "", {}, []):
        return "text/plain"
    return ""


def _endpoint_id(method: str, path: str, index: int) -> str:
    return f"req_{method.lower()}_{_safe_key(path, 'path')}_{index}"


def _openapi_endpoints(data: Dict[str, Any], source: Dict[str, Any]) -> List[Dict[str, Any]]:
    servers = data.get("servers") or []
    base_url = ""
    if servers and isinstance(servers, list):
        base_url = str((servers[0] or {}).get("url") or "")
    endpoints: List[Dict[str, Any]] = []
    for path, path_item in (data.get("paths") or {}).items():
        if not isinstance(path_item, dict):
            continue
        for method, operation in path_item.items():
            method_upper = str(method).upper()
            if method_upper not in HTTP_METHODS or not isinstance(operation, dict):
                continue
            headers: List[Dict[str, str]] = []
            query: List[Dict[str, str]] = []
            path_params: List[Dict[str, str]] = []
            for param in operation.get("parameters") or []:
                if not isinstance(param, dict):
                    continue
                item = {
                    "key": str(param.get("name") or ""),
                    "value": "",
                    "description": str(param.get("description") or ""),
                    "required": bool(param.get("required")),
                }
                if param.get("in") == "header":
                    headers.append(item)
                elif param.get("in") == "query":
                    query.append(item)
                elif param.get("in") == "path":
                    path_params.append(item)
            body: Dict[str, Any] = {"mode": "none"}
            request_body = operation.get("requestBody") or {}
            content = request_body.get("content") or {}
            if content:
                content_type, content_data = next(iter(content.items()))
                example = content_data.get("example")
                if example is None:
                    examples = content_data.get("examples") or {}
                    if examples:
                        example = next(iter(examples.values())).get("value")
                if example is None and "schema" in content_data:
                    example = {"TODO": "Completar body segun schema"}
                body = {
                    "mode": "raw",
                    "content_type": content_type,
                    "raw": json.dumps(example, ensure_ascii=False, indent=2) if not isinstance(example, str) else example,
                }
            responses = []
            for status, response in (operation.get("responses") or {}).items():
                responses.append({
                    "status": str(status),
                    "description": str((response or {}).get("description") or ""),
                    "source": source["name"],
                })
            auth = {"type": "inherit"}
            security = operation.get("security", data.get("security"))
            if security == []:
                auth = {"type": "noauth"}
            endpoint = {
                "id": _endpoint_id(method_upper, path, len(endpoints) + 1),
                "name": operation.get("summary") or operation.get("operationId") or f"{method_upper} {path}",
                "description": operation.get("description") or "",
                "method": method_upper,
                "protocol": "https" if str(base_url).startswith("https") else "",
                "base_url": base_url,
                "path": str(path),
                "path_params": path_params,
                "query_params": query,
                "headers": headers,
                "cookies": [],
                "auth": auth,
                "body": body,
                "content_type": body.get("content_type") or _infer_content_type(headers, body.get("raw")),
                "responses": responses,
                "variables": [],
                "dependencies": [],
                "status": "active",
                "source_refs": [_source_ref(source["name"], "openapi", f"paths.{path}.{method}")],
            }
            if method_upper == "GET" and body.get("mode") != "none":
                endpoint.setdefault("warnings", []).append("GET con body definido explicitamente por la documentacion.")
            endpoints.append(endpoint)
    return endpoints


def _postman_items(items: List[Dict[str, Any]], source: Dict[str, Any], folder: str = "") -> List[Dict[str, Any]]:
    endpoints: List[Dict[str, Any]] = []
    for item in items or []:
        if "item" in item:
            endpoints.extend(_postman_items(item.get("item") or [], source, item.get("name") or folder))
            continue
        request = item.get("request") or {}
        if isinstance(request, str):
            continue
        url = request.get("url") or {}
        raw_url = url.get("raw") if isinstance(url, dict) else str(url)
        parsed = urlparse(raw_url.replace("{{", "").replace("}}", ""))
        headers = [
            {"key": h.get("key", ""), "value": h.get("value", ""), "description": h.get("description", "")}
            for h in request.get("header") or []
            if isinstance(h, dict)
        ]
        query = [{"key": k, "value": v, "description": ""} for k, v in parse_qsl(parsed.query)]
        body_data = request.get("body") or {}
        body = {"mode": body_data.get("mode") or "none"}
        if body["mode"] == "raw":
            body.update({"raw": body_data.get("raw") or "", "content_type": _infer_content_type(headers, body_data.get("raw"))})
        elif body["mode"] in {"urlencoded", "formdata"}:
            body.update({"items": body_data.get(body["mode"]) or []})
        endpoint = {
            "id": _endpoint_id(request.get("method", "GET").upper(), parsed.path or raw_url, len(endpoints) + 1),
            "name": item.get("name") or raw_url,
            "description": request.get("description") or "",
            "method": request.get("method", "GET").upper(),
            "protocol": parsed.scheme,
            "base_url": f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else "",
            "path": parsed.path or raw_url,
            "path_params": [],
            "query_params": query,
            "headers": headers,
            "cookies": [],
            "auth": request.get("auth") or {"type": "inherit"},
            "body": body,
            "content_type": body.get("content_type") or _infer_content_type(headers, body.get("raw")),
            "responses": [],
            "variables": [],
            "dependencies": [],
            "status": "active",
            "folder": folder,
            "source_refs": [_source_ref(source["name"], "postman_collection", item.get("name", ""))],
        }
        endpoints.append(endpoint)
    return endpoints


def _curl_endpoints(text: str, source: Dict[str, Any]) -> List[Dict[str, Any]]:
    endpoints: List[Dict[str, Any]] = []
    chunks = re.split(r"(?=\bcurl\s+)", text, flags=re.I)
    for chunk in chunks:
        if not chunk.strip().lower().startswith("curl"):
            continue
        try:
            parts = shlex.split(chunk.strip(), posix=False)
        except Exception:
            parts = chunk.strip().split()
        method = "GET"
        headers: List[Dict[str, str]] = []
        body_raw = ""
        url = ""
        index = 1
        while index < len(parts):
            part = parts[index].strip("'\"")
            next_part = parts[index + 1].strip("'\"") if index + 1 < len(parts) else ""
            if part in {"-X", "--request"} and next_part:
                method = next_part.upper()
                index += 2
                continue
            if part in {"-H", "--header"} and next_part:
                key, _, value = next_part.partition(":")
                headers.append({"key": key.strip(), "value": value.strip(), "description": ""})
                index += 2
                continue
            if part in {"-d", "--data", "--data-raw", "--data-binary"} and next_part:
                body_raw = next_part
                if method == "GET":
                    method = "POST"
                index += 2
                continue
            if part.startswith("http"):
                url = part
            index += 1
        if not url:
            continue
        parsed = urlparse(url)
        body = {"mode": "none"} if not body_raw else {"mode": "raw", "raw": body_raw, "content_type": _infer_content_type(headers, body_raw)}
        endpoints.append({
            "id": _endpoint_id(method, parsed.path or url, len(endpoints) + 1),
            "name": f"{method} {parsed.path or url}",
            "description": "Request detectado desde cURL.",
            "method": method,
            "protocol": parsed.scheme,
            "base_url": f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else "",
            "path": parsed.path or url,
            "path_params": [],
            "query_params": [{"key": k, "value": v, "description": ""} for k, v in parse_qsl(parsed.query)],
            "headers": headers,
            "cookies": [],
            "auth": _auth_from_headers(headers),
            "body": body,
            "content_type": body.get("content_type") or _infer_content_type(headers, body.get("raw")),
            "responses": [],
            "variables": [],
            "dependencies": [],
            "status": "active",
            "source_refs": [_source_ref(source["name"], "curl")],
        })
    return endpoints


def _text_endpoints(text: str, source: Dict[str, Any]) -> List[Dict[str, Any]]:
    endpoints: List[Dict[str, Any]] = []
    pattern = re.compile(r"\b(GET|POST|PUT|PATCH|DELETE|HEAD|OPTIONS)\s+((?:https?://|/)[^\s)]+)", re.I)
    for match in pattern.finditer(text or ""):
        method = match.group(1).upper()
        raw_url = match.group(2).strip().rstrip(".,")
        parsed = urlparse(raw_url)
        endpoints.append({
            "id": _endpoint_id(method, parsed.path or raw_url, len(endpoints) + 1),
            "name": f"{method} {parsed.path or raw_url}",
            "description": "Endpoint detectado desde texto.",
            "method": method,
            "protocol": parsed.scheme,
            "base_url": f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else "",
            "path": parsed.path or raw_url,
            "path_params": [],
            "query_params": [{"key": k, "value": v, "description": ""} for k, v in parse_qsl(parsed.query)],
            "headers": [],
            "cookies": [],
            "auth": {"type": "inherit"},
            "body": {"mode": "none"},
            "content_type": "",
            "responses": [],
            "variables": [],
            "dependencies": [],
            "status": "active",
            "source_refs": [_source_ref(source["name"], "text", f"char {match.start()}")],
        })
    return endpoints


def _auth_from_headers(headers: List[Dict[str, str]]) -> Dict[str, Any]:
    for header in headers:
        key = header.get("key", "").lower()
        value = header.get("value", "")
        if key == "authorization":
            if value.lower().startswith("bearer"):
                return {"type": "bearer", "bearer": [{"key": "token", "value": "{{bearerToken}}", "type": "string"}]}
            if value.lower().startswith("basic"):
                return {"type": "basic"}
            return {"type": "apikey", "apikey": [{"key": "key", "value": "Authorization", "type": "string"}]}
        if "api" in key and "key" in key:
            return {"type": "apikey", "apikey": [{"key": "key", "value": header.get("key"), "type": "string"}]}
    return {"type": "inherit"}


def _extract_test_cases(text: str, source: Dict[str, Any]) -> List[Dict[str, Any]]:
    cases: List[Dict[str, Any]] = []
    blocks = re.split(r"(?im)(?=^\s*(?:caso|test case|tc|cp)[\s#:_-]*\d*)", text or "")
    for idx, block in enumerate(blocks):
        clean = block.strip()
        if len(clean) < 25:
            continue
        if not re.search(r"(?i)\b(caso|test case|prueba|resultado esperado|expected|paso|step)", clean):
            continue
        first_line = clean.splitlines()[0][:160]
        case_id_match = re.search(r"(?i)\b(?:TC|CP|CASO|TEST CASE)[\s#:_-]*([A-Za-z0-9_.-]+)", first_line)
        case_id = case_id_match.group(1) if case_id_match else f"CASE-{len(cases) + 1:03d}"
        expected = ""
        expected_match = re.search(r"(?is)(resultado esperado|expected result|expected)\s*[:\-]\s*(.+?)(?:\n\s*\n|$)", clean)
        if expected_match:
            expected = expected_match.group(2).strip()[:1000]
        cases.append({
            "id": f"case_{len(cases) + 1}",
            "case_id": case_id,
            "name": first_line.strip("# :-") or f"Caso {len(cases) + 1}",
            "objective": "",
            "description": clean[:2500],
            "preconditions": [],
            "input_data": {},
            "steps": [line.strip(" -\t") for line in clean.splitlines()[1:] if line.strip()][:40],
            "expected_result": expected,
            "postconditions": [],
            "priority": "",
            "test_type": _infer_case_type(clean),
            "dependencies": [],
            "created_or_modified_data": [],
            "related_request_ids": [],
            "source_refs": [_source_ref(source["name"], "test_case", first_line[:80])],
        })
    return cases


def _infer_case_type(text: str) -> str:
    normalized = (text or "").lower()
    for key in ("seguridad", "security", "autorizacion", "authorization", "negativo", "negative", "borde", "edge", "concurrencia", "idempotencia", "regresion"):
        if key in normalized:
            return key
    return "funcional"


def _extract_variables(sources: List[Dict[str, Any]], endpoints: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    variables: Dict[str, Dict[str, Any]] = {}
    warnings: List[Dict[str, Any]] = []

    def add_var(key: str, scope: str, source: str, value: str = "", sensitive: bool = False, reason: str = "") -> None:
        safe = _safe_key(key, "variable")
        if safe not in variables:
            variables[safe] = {
                "key": safe,
                "value": "" if sensitive else value,
                "scope": scope,
                "source": source,
                "sensitive": sensitive,
                "reason": reason,
                "enabled": True,
            }

    for endpoint in endpoints:
        if endpoint.get("base_url"):
            add_var("baseUrl", "environment", endpoint["source_refs"][0]["source"], endpoint["base_url"], False, "Base URL detectada")
            endpoint["base_url"] = "{{baseUrl}}"
        for header in endpoint.get("headers", []):
            if re.search(r"(?i)(authorization|token|api[-_]?key|secret|cookie)", header.get("key", "")):
                add_var(header.get("key", "secret"), "environment", endpoint["source_refs"][0]["source"], "", True, "Header sensible")
                header["value"] = "{{" + _safe_key(header.get("key", "secret")) + "}}"

    for source in sources:
        text = source.get("text", "")
        for name in re.findall(r"\{\{([A-Za-z0-9_.-]+)\}\}", text):
            add_var(name, "environment", source["name"], "", False, "Variable marcada en documentacion")
        for secret in SECRET_RE.finditer(text):
            warnings.append({
                "severity": "high",
                "type": "possible_secret",
                "message": "Se detecto un posible secreto. No se exporta su valor por defecto.",
                "source": source["name"],
                "sample": secret.group(0)[:80],
            })
    return list(variables.values()), warnings


def _associate(cases: List[Dict[str, Any]], endpoints: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    associations: List[Dict[str, Any]] = []
    for case in cases:
        case_blob = " ".join([case.get("name", ""), case.get("description", ""), case.get("expected_result", "")]).lower()
        best: Optional[Tuple[int, Dict[str, Any], List[str]]] = None
        for endpoint in endpoints:
            evidence: List[str] = []
            score = 0
            path_tokens = [token for token in re.split(r"[^a-zA-Z0-9]+", endpoint.get("path", "").lower()) if len(token) > 2]
            name_tokens = [token for token in re.split(r"[^a-zA-Z0-9]+", endpoint.get("name", "").lower()) if len(token) > 2]
            if endpoint.get("method", "").lower() in case_blob:
                score += 20
                evidence.append(f"Metodo {endpoint.get('method')} mencionado")
            matches = [token for token in set(path_tokens + name_tokens) if token in case_blob]
            if matches:
                score += min(60, len(matches) * 15)
                evidence.append("Coincidencias semanticas: " + ", ".join(matches[:6]))
            for response in endpoint.get("responses", []):
                status = str(response.get("status"))
                if status and status in case_blob:
                    score += 10
                    evidence.append(f"Status {status} mencionado")
            if not best or score > best[0]:
                best = (score, endpoint, evidence)
        if not best or best[0] < 15:
            associations.append({
                "case_id": case["id"],
                "endpoint_id": "",
                "confidence": "Sin coincidencia",
                "score": 0,
                "evidence": [],
                "explanation": "No se encontro evidencia suficiente para asociar este caso.",
                "source": case["source_refs"][0]["source"],
                "confirmed": False,
            })
        else:
            confidence = "Alta" if best[0] >= 70 else "Media" if best[0] >= 40 else "Baja"
            associations.append({
                "case_id": case["id"],
                "endpoint_id": best[1]["id"],
                "confidence": confidence,
                "score": best[0],
                "evidence": best[2],
                "explanation": f"Asociacion propuesta por coincidencias entre caso y request ({confidence}).",
                "source": case["source_refs"][0]["source"],
                "confirmed": confidence == "Alta",
            })
    return associations


def _case_blob(case: Dict[str, Any]) -> str:
    return " ".join([
        str(case.get("name", "")),
        str(case.get("description", "")),
        str(case.get("expected_result", "")),
        " ".join(str(step) for step in case.get("steps", [])),
    ]).lower()


def _generated_case(endpoint: Dict[str, Any], suffix: str, title: str, description: str, expected: str, case_number: int) -> Dict[str, Any]:
    method = endpoint.get("method", "GET")
    path = endpoint.get("path", endpoint.get("name", "endpoint"))
    return {
        "id": f"case_auto_{endpoint.get('id', case_number)}_{suffix}",
        "case_id": f"AUTO-{case_number:03d}",
        "name": f"{method} {path} - {title}",
        "objective": title,
        "description": description,
        "preconditions": [],
        "input_data": {},
        "steps": [
            f"Preparar request {method} {path}.",
            "Completar variables obligatorias con datos validos.",
            "Ejecutar request desde Postman.",
            "Validar status, estructura y mensaje de respuesta.",
        ],
        "expected_result": expected,
        "postconditions": [],
        "priority": "Media",
        "test_type": "funcional",
        "dependencies": [],
        "created_or_modified_data": [],
        "related_request_ids": [endpoint.get("id", "")],
        "source_refs": [_source_ref("generador_qa_senior", "generated_case", f"{method} {path}")],
        "generated": True,
    }


def _complete_qa_case_coverage(
    endpoints: List[Dict[str, Any]],
    test_cases: List[Dict[str, Any]],
    associations: List[Dict[str, Any]],
    compare_excel: bool = False,
) -> Tuple[List[Dict[str, Any]], Dict[str, List[Dict[str, str]]]]:
    case_by_id = {case["id"]: case for case in test_cases}
    cases_by_endpoint: Dict[str, List[Dict[str, Any]]] = {}
    for assoc in associations:
        if assoc.get("endpoint_id") and assoc.get("case_id") in case_by_id:
            cases_by_endpoint.setdefault(assoc["endpoint_id"], []).append(case_by_id[assoc["case_id"]])

    generated: List[Dict[str, Any]] = []
    adjustments = {"added": [], "extra": []}
    case_number = len(test_cases) + 1

    for endpoint in endpoints:
        endpoint_cases = cases_by_endpoint.get(endpoint.get("id", ""), [])
        blobs = " ".join(_case_blob(case) for case in endpoint_cases)
        method = endpoint.get("method", "GET")
        path = endpoint.get("path", endpoint.get("name", "endpoint"))

        needs = [
            (
                "ok",
                "camino feliz",
                f"Validar que el endpoint {method} {path} responda correctamente con datos validos.",
                "Respuesta exitosa y payload coherente con la documentacion.",
                not any(word in blobs for word in ("ok", "exito", "exitoso", "valido", "happy")),
            )
        ]
        if method in {"POST", "PUT", "PATCH"} or (endpoint.get("body") or {}).get("mode") not in {None, "none"}:
            needs.append((
                "required",
                "campos obligatorios",
                f"Enviar {method} {path} omitiendo un dato obligatorio o enviando formato invalido.",
                "La API rechaza la solicitud con error controlado y mensaje claro.",
                not any(word in blobs for word in ("obligatorio", "required", "invalido", "400", "validacion")),
            ))
        if endpoint.get("auth", {}).get("type") != "noauth":
            needs.append((
                "auth",
                "sin autorizacion",
                f"Ejecutar {method} {path} sin token, con token vencido o credencial invalida.",
                "La API responde 401 o 403 sin exponer informacion sensible.",
                not any(word in blobs for word in ("401", "403", "token", "autorizacion", "authorization")),
            ))
        if endpoint.get("path_params") or re.search(r"\{[^}]+\}|:[A-Za-z0-9_]+", path):
            needs.append((
                "not_found",
                "identificador inexistente",
                f"Ejecutar {method} {path} con identificador inexistente o mal formado.",
                "La API responde error controlado, idealmente 400 o 404.",
                not any(word in blobs for word in ("404", "inexistente", "no encontrado", "not found")),
            ))
        if method == "GET" and endpoint.get("query_params"):
            needs.append((
                "filters",
                "filtros y paginacion",
                f"Ejecutar {method} {path} combinando filtros, valores vacios y limites de paginacion.",
                "La API filtra correctamente o informa error de validacion cuando corresponde.",
                not any(word in blobs for word in ("filtro", "query", "paginacion", "pagina")),
            ))

        for suffix, title, description, expected, should_add in needs:
            if not should_add:
                continue
            case = _generated_case(endpoint, suffix, title, description, expected, case_number)
            generated.append(case)
            adjustments["added"].append({"endpoint": f"{method} {path}", "case": case["name"]})
            case_number += 1

    if compare_excel:
        for assoc in associations:
            if not assoc.get("endpoint_id") and assoc.get("case_id") in case_by_id:
                case = case_by_id[assoc["case_id"]]
                adjustments["extra"].append({"case": case.get("name", case.get("case_id", "")), "reason": "No se encontro endpoint relacionado."})

    if not compare_excel:
        adjustments = {"added": [], "extra": []}
    return generated, adjustments


def _detect_conflicts(endpoints: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    conflicts: List[Dict[str, Any]] = []
    seen: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for endpoint in endpoints:
        key = (endpoint.get("method", ""), endpoint.get("path", ""))
        current = seen.get(key)
        if current and current.get("body") != endpoint.get("body"):
            conflicts.append({
                "type": "operation_body_conflict",
                "message": f"Distintas fuentes describen body diferente para {key[0]} {key[1]}.",
                "values": [
                    {"source": current["source_refs"][0]["source"], "value": current.get("body")},
                    {"source": endpoint["source_refs"][0]["source"], "value": endpoint.get("body")},
                ],
                "proposed_priority": "OpenAPI > Documentacion API > Postman > Casos > Inferencia IA",
                "resolved": False,
            })
        seen.setdefault(key, endpoint)
    return conflicts


def build_intermediate_model(sources: List[Dict[str, Any]], manual_text: str = "") -> Dict[str, Any]:
    if manual_text.strip():
        sources.append({"name": "texto_pegado", "extension": ".txt", "text": manual_text, "size": len(manual_text), "warning": ""})
    endpoints: List[Dict[str, Any]] = []
    test_cases: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    for source in sources:
        source["kinds"] = _classify_source(source)
        if source.get("warning"):
            warnings.append({"severity": "medium", "type": "source_read", "message": source["warning"], "source": source["name"]})
        data = _try_load_structured(source.get("text", ""))
        if isinstance(data, dict) and ("openapi" in data or "swagger" in data or "paths" in data):
            endpoints.extend(_openapi_endpoints(data, source))
        if isinstance(data, dict) and data.get("info", {}).get("schema", "").endswith("collection/v2.1.0/collection.json"):
            endpoints.extend(_postman_items(data.get("item") or [], source))
        endpoints.extend(_curl_endpoints(source.get("text", ""), source))
        endpoints.extend(_text_endpoints(source.get("text", ""), source))
        test_cases.extend(_extract_test_cases(source.get("text", ""), source))
    variables, secret_warnings = _extract_variables(sources, endpoints)
    warnings.extend(secret_warnings)
    associations = _associate(test_cases, endpoints)
    generated_cases, case_adjustments = _complete_qa_case_coverage(
        endpoints,
        test_cases,
        associations,
        any(source.get("is_cases_file") for source in sources),
    )
    if generated_cases:
        test_cases.extend(generated_cases)
        associations = _associate(test_cases, endpoints)
    endpoint_ids_with_case = {a["endpoint_id"] for a in associations if a.get("endpoint_id")}
    for endpoint in endpoints:
        if endpoint["id"] not in endpoint_ids_with_case:
            warnings.append({
                "severity": "low",
                "type": "endpoint_without_case",
                "message": f"Endpoint sin caso asociado: {endpoint['method']} {endpoint['path']}",
                "source": endpoint["source_refs"][0]["source"],
            })
    for assoc in associations:
        if not assoc.get("endpoint_id"):
            warnings.append({
                "severity": "medium",
                "type": "case_without_endpoint",
                "message": f"Caso sin endpoint asociado: {assoc['case_id']}",
                "source": assoc.get("source", ""),
            })
    return {
        "version": "postman_intermediate_v1",
        "sources": [{k: v for k, v in source.items() if k != "text"} for source in sources],
        "endpoints": endpoints,
        "test_cases": test_cases,
        "associations": associations,
        "variables": variables,
        "dependencies": [],
        "warnings": warnings,
        "conflicts": _detect_conflicts(endpoints),
        "case_adjustments": case_adjustments,
        "folders": ["Endpoints", "Casos de prueba"],
        "validation": {},
    }


def _postman_url(endpoint: Dict[str, Any]) -> Dict[str, Any]:
    base = endpoint.get("base_url", "")
    path = endpoint.get("path", "")
    raw = (base.rstrip("/") + "/" + path.lstrip("/")).strip("/") if base else path
    if base.startswith("{{"):
        raw = base.rstrip("/") + "/" + path.lstrip("/")
    query = [
        {"key": item.get("key", ""), "value": item.get("value", ""), "description": item.get("description", "")}
        for item in endpoint.get("query_params", [])
    ]
    return {"raw": raw or "{{baseUrl}}/", "host": [base or "{{baseUrl}}"], "path": [part for part in path.strip("/").split("/") if part], "query": query}


def _postman_body(body: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    mode = (body or {}).get("mode", "none")
    if mode == "none":
        return None
    if mode == "raw":
        return {
            "mode": "raw",
            "raw": body.get("raw", ""),
            "options": {"raw": {"language": "json" if "json" in body.get("content_type", "") else "text"}},
        }
    if mode in {"urlencoded", "formdata"}:
        return {"mode": mode, mode: body.get("items") or []}
    if mode == "file":
        return {"mode": "file", "file": {"src": body.get("src", "")}}
    return {"mode": "raw", "raw": str(body.get("raw", ""))}


def _tests_for_endpoint(endpoint: Dict[str, Any], associations: List[Dict[str, Any]], cases: List[Dict[str, Any]]) -> str:
    related = [a for a in associations if a.get("endpoint_id") == endpoint.get("id")]
    statuses = {
        str(resp.get("status"))
        for resp in endpoint.get("responses", [])
        if str(resp.get("status", "")).isdigit()
    }
    lines = [
        "pm.test('La respuesta tiene un status documentado o aceptable', function () {",
    ]
    if statuses:
        lines.append(f"  pm.expect([{', '.join(sorted(statuses))}]).to.include(pm.response.code);")
    else:
        lines.append("  pm.expect(pm.response.code).to.be.within(100, 599);")
    lines.append("});")
    if endpoint.get("content_type"):
        lines.extend([
            "",
            "pm.test('Content-Type coherente si viene informado', function () {",
            "  const contentType = pm.response.headers.get('Content-Type') || '';",
            f"  pm.expect(contentType.toLowerCase()).to.include('{endpoint['content_type'].split(';')[0].lower()}');",
            "});",
        ])
    if related:
        case_names = [case.get("name") for case in cases if case.get("id") in {a["case_id"] for a in related}]
        lines.extend(["", f"console.log('Casos relacionados: {json.dumps(case_names, ensure_ascii=False)}');"])
    lines.extend(["", "// TODO_ASSERTION: Agregar validaciones especificas confirmadas por QA."])
    return "\n".join(lines)


def build_collection(model: Dict[str, Any]) -> Dict[str, Any]:
    associations = model.get("associations", [])
    cases = model.get("test_cases", [])
    items: List[Dict[str, Any]] = []
    for endpoint in model.get("endpoints", []):
        if endpoint.get("status") in {"disabled", "blocked"}:
            continue
        request: Dict[str, Any] = {
            "method": endpoint.get("method", "GET"),
            "header": endpoint.get("headers", []),
            "url": _postman_url(endpoint),
            "description": endpoint.get("description", ""),
        }
        body = _postman_body(endpoint.get("body") or {})
        if body:
            request["body"] = body
        auth = endpoint.get("auth") or {"type": "inherit"}
        if auth.get("type") and auth.get("type") != "inherit":
            request["auth"] = auth
        item = {
            "name": endpoint.get("name") or f"{endpoint.get('method')} {endpoint.get('path')}",
            "request": request,
            "event": [{"listen": "test", "script": {"type": "text/javascript", "exec": _tests_for_endpoint(endpoint, associations, cases).splitlines()}}],
        }
        items.append(item)
    return {
        "info": {
            "name": "Coleccion QA generada",
            "schema": POSTMAN_SCHEMA,
            "description": "Generada por QA Doc Analyzer desde documentacion, casos y fuentes adjuntas.",
        },
        "item": [{"name": model.get("folders", ["Endpoints"])[0] or "Endpoints", "item": items}],
        "variable": [
            {"key": var["key"], "value": var.get("value", ""), "type": "string"}
            for var in model.get("variables", [])
            if var.get("scope") == "collection"
        ],
    }


def build_environment(model: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "name": "QA Environment",
        "values": [
            {
                "key": var["key"],
                "value": "" if var.get("sensitive") else var.get("value", ""),
                "type": "secret" if var.get("sensitive") else "default",
                "enabled": bool(var.get("enabled", True)),
            }
            for var in model.get("variables", [])
            if var.get("scope") in {"environment", ""}
        ],
        "_postman_variable_scope": "environment",
        "_postman_exported_using": "QA Doc Analyzer",
    }


def build_traceability(model: Dict[str, Any]) -> str:
    endpoint_by_id = {item["id"]: item for item in model.get("endpoints", [])}
    case_by_id = {item["id"]: item for item in model.get("test_cases", [])}
    lines = ["# Informe de trazabilidad", "", "| Caso | Request | Confianza | Evidencia |", "|---|---|---|---|"]
    for assoc in model.get("associations", []):
        case = case_by_id.get(assoc.get("case_id"), {})
        endpoint = endpoint_by_id.get(assoc.get("endpoint_id"), {})
        req = f"{endpoint.get('method', '')} {endpoint.get('path', '')}".strip() or "Sin coincidencia"
        lines.append(
            f"| {case.get('case_id', assoc.get('case_id'))} - {case.get('name', '')} | "
            f"{req} | {assoc.get('confidence')} | {'; '.join(assoc.get('evidence') or [])} |"
        )
    return "\n".join(lines)


def build_readme(model: Dict[str, Any]) -> str:
    warnings = model.get("warnings", [])
    conflicts = model.get("conflicts", [])
    return "\n".join([
        "# Coleccion Postman generada",
        "",
        "## Importacion",
        "1. Importar `collection.json` en Postman.",
        "2. Importar `environment.json` si fue generado.",
        "3. Revisar variables vacias antes de ejecutar.",
        "",
        "## Autenticacion",
        "La autenticacion fue conservada solamente cuando aparecio evidencia en las fuentes. Requests sin evidencia heredan la configuracion o quedan sin auth especifica.",
        "",
        "## Orden de ejecucion",
        "Ejecutar requests individuales o por carpeta. Revisar la trazabilidad para flujos multi-step y dependencias pendientes.",
        "",
        "## Variables pendientes",
        *[f"- `{var['key']}` ({var.get('scope', 'environment')}): {var.get('reason', '')}" for var in model.get("variables", [])],
        "",
        "## Advertencias",
        *(f"- [{w.get('severity')}] {w.get('message')} ({w.get('source', '')})" for w in warnings),
        "",
        "## Conflictos",
        *(f"- {c.get('message')} - resuelto: {c.get('resolved', False)}" for c in conflicts),
    ])


def validate_model(model: Dict[str, Any]) -> Dict[str, Any]:
    errors: List[str] = []
    warnings: List[str] = []
    for endpoint in model.get("endpoints", []):
        if endpoint.get("method") not in HTTP_METHODS:
            errors.append(f"Metodo invalido en {endpoint.get('name')}.")
        if not endpoint.get("path"):
            errors.append(f"Request sin path/url: {endpoint.get('name')}.")
        if endpoint.get("method") == "GET" and (endpoint.get("body") or {}).get("mode") not in {None, "none"}:
            warnings.append(f"GET con body en {endpoint.get('name')}; revisar antes de exportar.")
        content_type = endpoint.get("content_type") or ""
        body = endpoint.get("body") or {}
        if body.get("mode") == "raw" and "json" in content_type:
            try:
                json.loads(body.get("raw") or "{}")
            except Exception:
                warnings.append(f"Body JSON no parseable en {endpoint.get('name')}.")
    declared_vars = {var["key"] for var in model.get("variables", [])}
    used_vars = set(re.findall(r"\{\{([A-Za-z0-9_.-]+)\}\}", json.dumps(model, ensure_ascii=False)))
    missing_vars = sorted(used_vars - declared_vars)
    if missing_vars:
        warnings.append("Variables usadas no declaradas: " + ", ".join(missing_vars))
    if any(w.get("type") == "possible_secret" for w in model.get("warnings", [])):
        warnings.append("Hay posibles secretos detectados; confirmar antes de compartir la coleccion.")
    try:
        json.dumps(build_collection(model))
        json.dumps(build_environment(model))
    except Exception as exc:
        errors.append(f"No se pudo serializar Postman: {type(exc).__name__}.")
    return {"valid": not errors, "errors": errors, "warnings": warnings}


def _public(doc: Dict[str, Any]) -> Dict[str, Any]:
    public = dict(doc)
    public["id"] = str(public.pop("_id"))
    return public


def _require_qa(user: Dict[str, Any]) -> None:
    if user.get("role") != "qa":
        raise HTTPException(status_code=403, detail="Postman disponible solo para QA.")


async def _get_draft(draft_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    doc = await _db()[COLLECTION].find_one({"_id": _oid(draft_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Draft no encontrado.")
    if doc.get("created_by") != user.get("username"):
        raise HTTPException(status_code=403, detail="No podes acceder a este draft.")
    return doc


@router.post("/analyze")
async def analyze_postman_sources(
    files: Optional[List[UploadFile]] = File(default=None),
    test_cases_file: Optional[UploadFile] = File(default=None),
    manual_text: str = Form(default=""),
    project_name: str = Form(default=""),
    user: Dict[str, Any] = Depends(current_user),
):
    _require_qa(user)
    sources = [await _read_upload(file) for file in (files or [])]
    case_source = None
    if test_cases_file and test_cases_file.filename:
        case_source = await _read_upload(test_cases_file)
        case_source["is_cases_file"] = True
        sources.append(case_source)
    if not sources and not manual_text.strip():
        raise HTTPException(status_code=400, detail="Carga al menos un archivo o texto.")
    model = build_intermediate_model(sources, manual_text)
    if case_source:
        model["cases_source"] = {
            "name": case_source["name"],
            "detected_cases": len([case for case in model["test_cases"] if case.get("source_refs", [{}])[0].get("source") == case_source["name"]]),
        }
    model["validation"] = validate_model(model)
    now = _now()
    doc = {
        "project_name": _clean(project_name, 180) or "Proyecto API",
        "model": model,
        "created_by": user["username"],
        "created_at": now,
        "updated_at": now,
    }
    result = await _db()[COLLECTION].insert_one(doc)
    doc["_id"] = result.inserted_id
    await record_activity(
        user,
        "Generacion Postman",
        "postman",
        f"Analizo fuentes para Postman: {doc['project_name']}",
        {
            "draft_id": str(result.inserted_id),
            "fuentes": [source["name"] for source in sources if not source.get("is_cases_file")],
            "archivo_casos": case_source["name"] if case_source else "",
            "endpoints": len(model["endpoints"]),
            "casos": len(model["test_cases"]),
            "warnings": len(model["warnings"]),
        },
    )
    return {"draft": _public(doc)}


@router.get("/drafts")
async def list_drafts(user: Dict[str, Any] = Depends(current_user)):
    _require_qa(user)
    query = {"created_by": user["username"]}
    docs = await _db()[COLLECTION].find(query).sort("created_at", DESCENDING).to_list(100)
    return {"drafts": [_public(doc) for doc in docs]}


@router.get("/drafts/{draft_id}")
async def get_draft(draft_id: str, user: Dict[str, Any] = Depends(current_user)):
    _require_qa(user)
    return {"draft": _public(await _get_draft(draft_id, user))}


@router.put("/drafts/{draft_id}")
async def update_draft(draft_id: str, payload: DraftUpdate, user: Dict[str, Any] = Depends(current_user)):
    _require_qa(user)
    doc = await _get_draft(draft_id, user)
    model = dict(doc.get("model") or {})
    update_data = payload.model_dump(exclude_none=True)
    model.update(update_data)
    model["validation"] = validate_model(model)
    await _db()[COLLECTION].update_one({"_id": doc["_id"]}, {"$set": {"model": model, "updated_at": _now()}})
    doc["model"] = model
    await record_activity(
        user,
        "Revision Postman",
        "postman",
        f"Actualizo revision de Postman: {doc.get('project_name', 'Proyecto API')}",
        {
            "draft_id": draft_id,
            "project_name": doc.get("project_name"),
            "endpoints": len(model.get("endpoints", [])),
            "casos": len(model.get("test_cases", [])),
            "asociaciones": len([item for item in model.get("associations", []) if item.get("endpoint_id")]),
        },
    )
    return {"draft": _public(doc)}


def _file_payload(kind: str, model: Dict[str, Any]) -> Tuple[str, bytes, str]:
    if kind == "collection":
        return "collection.json", json.dumps(build_collection(model), ensure_ascii=False, indent=2).encode("utf-8"), "application/json"
    if kind == "environment":
        return "environment.json", json.dumps(build_environment(model), ensure_ascii=False, indent=2).encode("utf-8"), "application/json"
    if kind == "readme":
        return "README.md", build_readme(model).encode("utf-8"), "text/markdown"
    if kind == "traceability":
        return "traceability.md", build_traceability(model).encode("utf-8"), "text/markdown"
    raise HTTPException(status_code=404, detail="Archivo no soportado.")


@router.get("/drafts/{draft_id}/download/{kind}")
async def download_generated(kind: str, draft_id: str, user: Dict[str, Any] = Depends(current_user)):
    _require_qa(user)
    doc = await _get_draft(draft_id, user)
    model = doc.get("model") or {}
    label = "ZIP completo" if kind == "zip" else kind
    await record_activity(
        user,
        "Descarga Postman",
        "postman",
        f"Descargo {label} de Postman: {doc.get('project_name', 'Proyecto API')}",
        {
            "draft_id": draft_id,
            "project_name": doc.get("project_name"),
            "tipo_descarga": kind,
            "endpoints": len(model.get("endpoints", [])),
            "casos": len(model.get("test_cases", [])),
        },
    )
    if kind == "zip":
        output = BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
            for item_kind in ("collection", "environment", "readme", "traceability"):
                filename, content, _ = _file_payload(item_kind, model)
                zf.writestr(filename, content)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{OUTPUT_NAME}.zip"'},
        )
    filename, content, media_type = _file_payload(kind, model)
    if media_type.startswith("text"):
        return PlainTextResponse(content.decode("utf-8"), media_type=media_type, headers={"Content-Disposition": f'attachment; filename="{filename}"'})
    return Response(content, media_type=media_type, headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.post("/drafts/{draft_id}/validate")
async def validate_draft(draft_id: str, user: Dict[str, Any] = Depends(current_user)):
    _require_qa(user)
    doc = await _get_draft(draft_id, user)
    validation = validate_model(doc.get("model") or {})
    await _db()[COLLECTION].update_one({"_id": doc["_id"]}, {"$set": {"model.validation": validation, "updated_at": _now()}})
    await record_activity(
        user,
        "Validacion Postman",
        "postman",
        f"Valido coleccion Postman: {doc.get('project_name', 'Proyecto API')}",
        {
            "draft_id": draft_id,
            "project_name": doc.get("project_name"),
            "valid": validation.get("valid"),
            "errores": len(validation.get("errors", [])),
            "advertencias": len(validation.get("warnings", [])),
        },
    )
    return {"validation": validation}
